"""
build_user_inputs_sheet.py
==========================
Two jobs on ModelBaseline5Sectorsand1Regions.xlsx:

  1. Fix Baseline_Input rows 24-48 (manual annual growth-factor display table):
     Replace stale hardcoded values with OFFSET formulas that always mirror the
     intermediate gY / gN growth factors in rows 142-146 / 163-167.

  2. Create / refresh 'User_Inputs' sheet.
     - EDITABLE rows  (yellow): the truly user-specified time series.
       Corresponding Baseline_Input year-cells become =User_Inputs!XNN refs.
     - DISPLAY rows   (grey)  : rows that are formula-computed in Baseline_Input
       (milestone interpolations, IO_Data pull-throughs, etc.).
       User_Inputs shows =Baseline_Input!XNN for reference only; no re-linking.

Idempotent: safe to re-run at any time.

Run from the repo root:
    python scripts/maintenance/build_user_inputs_sheet.py
"""

from pathlib import Path
from tempfile import NamedTemporaryFile
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font

REPO_ROOT   = Path(__file__).resolve().parent.parent.parent
WORKBOOK    = REPO_ROOT / "ExcelFiles" / "ModelBaseline5Sectorsand1Regions.xlsx"
SOURCE_WORKBOOK = REPO_ROOT / "ExcelFiles" / "ModelSimulationandCalibration5Sectorsand1Regions.xlsx"
SHEET_INPUT = "Baseline_Input"
SHEET_UI    = "User_Inputs"

BASE_YEAR  = 2025
FINAL_YEAR = 2050

# ── column helpers ──────────────────────────────────────────────────────────
def year_col(year: int) -> int:
    return 3 + (year - BASE_YEAR)          # col C = 2025

def col_l(year: int) -> str:
    return get_column_letter(year_col(year))

Y_START = year_col(BASE_YEAR)              # 3
Y_END   = year_col(FINAL_YEAR)             # 28

SECTORS = ["Primary", "Fossil", "Renewables", "Secondary", "Tertiary"]

# ── Rows in Baseline_Input that are EDITABLE by the user ────────────────────
#   (plain numbers; all others are formula-computed from milestones / IO_Data)
EDITABLE_ROWS = {
    238: "Total GDP growth rate (%)",
    242: "Primary VA share",
    243: "Fossil VA share",
    244: "Renewables VA share",
    245: "Secondary VA share",
    246: "Tertiary VA share",
    253: "Total employment growth rate (%)",
    257: "Primary employment share",
    258: "Fossil employment share",
    259: "Renewables employment share",
    260: "Secondary employment share",
    261: "Tertiary employment share",
    268: "Primary emissions (index)",
    269: "Fossil emissions (index)",
    270: "Renewables emissions (index)",
    271: "Secondary emissions (index)",
    272: "Tertiary emissions (index)",
    273: "Fossil sector output (index)",
    274: "Fossil exports (index)",
    279: "Primary public capital (index)",
    280: "Fossil public capital (index)",
    281: "Renewables public capital (index)",
    282: "Secondary public capital (index)",
    283: "Tertiary public capital (index)",
}

# Rows to show READ-ONLY in User_Inputs (formula-computed in Baseline_Input)
DISPLAY_ROWS_GVA = {}
DISPLAY_ROWS_EMP = {}
DISPLAY_ROWS_EMIS = {}

# ── styles ──────────────────────────────────────────────────────────────────
FILL_HEADER  = PatternFill("solid", fgColor="4472C4")
FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")
FILL_INPUT   = PatternFill("solid", fgColor="FFFF99")   # yellow — edit here
FILL_DISPLAY = PatternFill("solid", fgColor="F2F2F2")   # grey  — read-only

def _fill(ws, row, c_start, c_end, fill):
    for c in range(c_start, c_end + 1):
        ws.cell(row=row, column=c).fill = fill

# ── value helpers ────────────────────────────────────────────────────────────
UI_REF = f"='{SHEET_UI}'!"
BI_REF = f"='{SHEET_INPUT}'!"

def read_series_val(ws_val, row: int):
    """Read year-column values (data_only workbook)."""
    return [ws_val.cell(row=row, column=c).value for c in range(Y_START, Y_END + 1)]

def read_series_ui(ws_val, row: int):
    """Try to read a series from User_Inputs (data_only); fall back to None list."""
    if SHEET_UI not in ws_val.sheetnames:
        return [None] * (Y_END - Y_START + 1)
    ws = ws_val[SHEET_UI]
    return [ws.cell(row=row, column=c).value for c in range(Y_START, Y_END + 1)]

def resolve_editable_series(ws_bi_fml, ws_bi_val, wb_val_wb, bi_row: int, ws_fallback=None):
    """Return the current user values for an editable row, handling re-runs."""
    sample = ws_bi_fml.cell(row=bi_row, column=Y_START).value
    if isinstance(sample, str) and sample.startswith(UI_REF):
        # Already linked to User_Inputs — read from the UI sheet in the value wb
        # parse column / row from =User_Inputs!D12 style
        ref = sample[len(UI_REF):]          # e.g. "D12"
        ui_col = column_index_from_string(ref[:-len(ref.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))])
        ui_row = int(ref.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        ws_ui_val = wb_val_wb[SHEET_UI] if SHEET_UI in wb_val_wb.sheetnames else None
        vals = [ws_ui_val.cell(row=ui_row, column=c).value
                for c in range(Y_START, Y_END + 1)] if ws_ui_val is not None else []
        # Fall back to bi_val if UI values are missing
        if not vals or all(v is None for v in vals):
            vals = read_series_val(ws_bi_val, bi_row)
            if ws_fallback is not None and all(v is None for v in vals):
                return read_series_val(ws_fallback, bi_row)
            return vals
        return vals
    vals = read_series_val(ws_bi_val, bi_row)
    if ws_fallback is not None and all(v is None for v in vals):
        return read_series_val(ws_fallback, bi_row)
    return vals

def write_series_vals(ws, row, label, values, sec_idx=None):
    ws.cell(row=row, column=1).value = label
    if sec_idx is not None:
        ws.cell(row=row, column=2).value = sec_idx
    for i, v in enumerate(values):
        ws.cell(row=row, column=Y_START + i).value = v

def write_series_display(ws_ui, row, label, bi_row, sec_idx=None):
    """Write read-only =Baseline_Input!XNN references."""
    ws_ui.cell(row=row, column=1).value = label
    if sec_idx is not None:
        ws_ui.cell(row=row, column=2).value = sec_idx
    for c in range(Y_START, Y_END + 1):
        col = get_column_letter(c)
        ws_ui.cell(row=row, column=c).value = f"='{SHEET_INPUT}'!{col}{bi_row}"

def write_year_row(ws, row):
    ws.cell(row=row, column=1).value = "Year"
    for y in range(BASE_YEAR, FINAL_YEAR + 1):
        ws.cell(row=row, column=year_col(y)).value = y
    _fill(ws, row, 1, Y_END, FILL_DISPLAY)

def sum_check(ws, row, start, n):
    ws.cell(row=row, column=1).value = "Sum check (should be 1.0)"
    for c in range(Y_START, Y_END + 1):
        col = get_column_letter(c)
        ws.cell(row=row, column=c).value = f"=SUM({col}{start}:{col}{start+n-1})"
    _fill(ws, row, 1, Y_END, FILL_DISPLAY)

def link_to_ui(ws_bi, bi_row, ui_row):
    for c in range(Y_START, Y_END + 1):
        col = get_column_letter(c)
        ws_bi.cell(row=bi_row, column=c).value = f"='{SHEET_UI}'!{col}{ui_row}"


def sanitize_excel_package(workbook_path: Path):
    """Remove stale package artifacts that can trigger Excel repair mode.

    openpyxl can preserve a stale calcChain and an old workbook recovery flag.
    Excel will rebuild the calc chain itself, so dropping those artifacts is safe.
    """
    rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    ct_ns = {"ct": "http://schemas.openxmlformats.org/package/2006/content-types"}
    wb_ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    ET.register_namespace("", rel_ns["rel"])
    ET.register_namespace("", ct_ns["ct"])
    ET.register_namespace("", wb_ns["main"])

    with ZipFile(workbook_path, "r") as src, NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = Path(tmp.name)
        with ZipFile(tmp_path, "w", compression=ZIP_DEFLATED) as dst:
            for item in src.infolist():
                if item.filename == "xl/calcChain.xml":
                    continue

                data = src.read(item.filename)

                if item.filename == "xl/_rels/workbook.xml.rels":
                    root = ET.fromstring(data)
                    for rel in list(root):
                        if rel.attrib.get("Type", "").endswith("/calcChain"):
                            root.remove(rel)
                            continue
                        target = rel.attrib.get("Target")
                        if target == "/xl/styles.xml":
                            rel.set("Target", "styles.xml")
                        elif target == "/xl/sharedStrings.xml":
                            rel.set("Target", "sharedStrings.xml")
                        elif target == "/xl/theme/theme1.xml":
                            rel.set("Target", "theme/theme1.xml")
                        elif isinstance(target, str) and target.startswith("/xl/worksheets/"):
                            rel.set("Target", target[len("/xl/"):])
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                elif item.filename == "[Content_Types].xml":
                    root = ET.fromstring(data)
                    for override in list(root):
                        if override.attrib.get("PartName") == "/xl/calcChain.xml":
                            root.remove(override)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                elif item.filename == "xl/workbook.xml":
                    root = ET.fromstring(data)
                    recovery = root.find("main:fileRecoveryPr", wb_ns)
                    if recovery is not None:
                        root.remove(recovery)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)

                dst.writestr(item, data)

    workbook_path.unlink()
    tmp_path.replace(workbook_path)


def show_only_primary_baseline_blocks(ws_bi):
    """Keep Baseline_Input focused on the first two user-facing blocks.

    Baseline_calc still reads the hidden helper rows below, but the visible sheet
    should start with the VA/GVA block and then the employment block.
    """
    for row in range(1, 235):
        ws_bi.row_dimensions[row].hidden = True
    for row in range(235, 263):
        ws_bi.row_dimensions[row].hidden = False
    for row in range(263, ws_bi.max_row + 1):
        ws_bi.row_dimensions[row].hidden = True

    ws_bi.cell(row=235, column=1).value = "BLOCK 1 - VA SHARES AND TOTAL GVA GROWTH RATE"
    ws_bi.cell(row=236, column=1).value = (
        "Edit only the yellow cells in this block. These annual VA shares and the total "
        "GVA growth rate feed the hidden helper rows consumed by Baseline_calc."
    )
    ws_bi.cell(row=237, column=1).value = "Year"
    ws_bi.cell(row=238, column=1).value = "Total GVA growth rate (%)"
    ws_bi.cell(row=239, column=1).value = "Total GVA growth factor"
    ws_bi.cell(row=240, column=1).value = "Total GVA level (computed)"
    ws_bi.cell(row=241, column=1).value = "VA shares by economic activity"
    ws_bi.cell(row=247, column=1).value = "VA share sum check"

    ws_bi.cell(row=250, column=1).value = "BLOCK 2 - EMPLOYMENT SHARES AND TOTAL EMPLOYMENT GROWTH RATE"
    ws_bi.cell(row=251, column=1).value = (
        "Edit only the yellow cells in this block. These annual employment shares and "
        "the total employment growth rate feed the hidden helper rows consumed by Baseline_calc."
    )
    ws_bi.cell(row=252, column=1).value = "Year"
    ws_bi.cell(row=253, column=1).value = "Total employment growth rate (%)"
    ws_bi.cell(row=254, column=1).value = "Total employment growth factor"
    ws_bi.cell(row=255, column=1).value = "Total employment level (computed)"
    ws_bi.cell(row=256, column=1).value = "Employment shares by economic activity"
    ws_bi.cell(row=262, column=1).value = "Employment share sum check"

    # Open the worksheet on the first visible editable block.
    ws_bi.sheet_view.topLeftCell = "A235"
    ws_bi.sheet_view.selection[0].activeCell = "C238"
    ws_bi.sheet_view.selection[0].sqref = "C238"

# ── main ────────────────────────────────────────────────────────────────────
def main():
    wb_fml = openpyxl.load_workbook(str(WORKBOOK))
    wb_val = openpyxl.load_workbook(str(WORKBOOK), data_only=True)
    wb_src_val = openpyxl.load_workbook(str(SOURCE_WORKBOOK), data_only=True) if SOURCE_WORKBOOK.exists() else None
    ws_bi  = wb_fml[SHEET_INPUT]
    ws_biv = wb_val[SHEET_INPUT]
    ws_uiv = wb_val[SHEET_UI] if SHEET_UI in wb_val.sheetnames else None
    ws_src_biv = wb_src_val[SHEET_INPUT] if wb_src_val is not None and SHEET_INPUT in wb_src_val.sheetnames else None

    # ── Recreate User_Inputs ─────────────────────────────────────────────────
    if SHEET_UI in wb_fml.sheetnames:
        del wb_fml[SHEET_UI]
    ws_ui = wb_fml.create_sheet(SHEET_UI, 1)

    R = 1  # row pointer

    # title
    ws_ui.cell(row=R, column=1).value = "USER INPUTS  —  Baseline Scenario"
    ws_ui.cell(row=R, column=1).fill  = FILL_HEADER
    ws_ui.cell(row=R, column=1).font  = Font(bold=True, color="FFFFFF")
    ws_ui.merge_cells(start_row=R, start_column=1, end_row=R, end_column=Y_END)
    R += 2

    # ── Config ───────────────────────────────────────────────────────────────
    ws_ui.cell(row=R, column=1).value = "=== CONFIGURATION ==="
    ws_ui.cell(row=R, column=1).fill  = FILL_SECTION
    ws_ui.cell(row=R, column=1).font  = Font(bold=True)
    R += 1
    cfg = [
        ("Base year",                          3, 2),
        ("Final year",                         4, 2),
        ("Use milestones  (0=annual, 1=milestone table)", 5, 2),
        ("Sector share mode  (0=user annual, 1=linear)", 6, 2),
        ("GVA growth input mode  (0=factor, 1=percent)", 7, 2),
    ]
    cfg_links = {}
    for label, bi_r, bi_c in cfg:
        val = ws_biv.cell(row=bi_r, column=bi_c).value
        ws_ui.cell(row=R, column=1).value = label
        ws_ui.cell(row=R, column=3).value = val
        ws_ui.cell(row=R, column=3).fill  = FILL_INPUT
        cfg_links[(bi_r, bi_c)] = (R, 3)
        R += 1
    R += 1

    # ── helper: one editable row ─────────────────────────────────────────────
    def add_editable(bi_row, label, sec_idx=None):
        nonlocal R
        vals = resolve_editable_series(ws_bi, ws_biv, wb_val, bi_row, ws_fallback=ws_src_biv)  # noqa: F821 (wb_val defined in main)
        write_series_vals(ws_ui, R, label, vals, sec_idx=sec_idx)
        _fill(ws_ui, R, Y_START, Y_END, FILL_INPUT)
        ui_row = R
        R += 1
        return ui_row

    def add_display(bi_row, label, sec_idx=None):
        nonlocal R
        write_series_display(ws_ui, R, label, bi_row, sec_idx=sec_idx)
        _fill(ws_ui, R, Y_START, Y_END, FILL_DISPLAY)
        R += 1

    # ── Section 1: GVA ───────────────────────────────────────────────────────
    ws_ui.cell(row=R, column=1).value = "=== SECTION 1:  TOTAL GVA GROWTH & VA SHARES ==="
    ws_ui.cell(row=R, column=1).fill  = FILL_SECTION
    ws_ui.cell(row=R, column=1).font  = Font(bold=True)
    R += 1
    write_year_row(ws_ui, R); R += 1

    ui_238 = add_editable(238, "Total GVA growth rate  (%)  [EDIT HERE]")
    R += 1  # blank
    ws_ui.cell(row=R, column=1).value = "VA shares by economic activity (fractions, sum to 1.0)  [EDIT HERE]"
    R += 1
    SEC_Y_START = R
    gva_share_links = {}
    for i, bi_r in enumerate(range(242, 247)):
        gva_share_links[bi_r] = add_editable(bi_r, SECTORS[i], sec_idx=i + 1)
    sum_check(ws_ui, R, SEC_Y_START, 5)
    R += 2

    # ── Section 2: Employment ────────────────────────────────────────────────
    ws_ui.cell(row=R, column=1).value = "=== SECTION 2:  EMPLOYMENT GROWTH & SECTOR SHARES ==="
    ws_ui.cell(row=R, column=1).fill  = FILL_SECTION
    ws_ui.cell(row=R, column=1).font  = Font(bold=True)
    R += 1
    write_year_row(ws_ui, R); R += 1

    ui_253 = add_editable(253, "Total employment growth rate  (%)  [EDIT HERE]")
    R += 1
    ws_ui.cell(row=R, column=1).value = "Employment shares (fractions, sum to 1.0)  [EDIT HERE]"
    R += 1
    SEC_N_START = R
    emp_share_links = {}
    for i, bi_r in enumerate(range(257, 262)):
        emp_share_links[bi_r] = add_editable(bi_r, SECTORS[i], sec_idx=i + 1)
    sum_check(ws_ui, R, SEC_N_START, 5)
    R += 2

    # ── Section 3: Emissions ─────────────────────────────────────────────────
    ws_ui.cell(row=R, column=1).value = "=== SECTION 3:  EMISSIONS & FOSSIL OUTPUT ==="
    ws_ui.cell(row=R, column=1).fill  = FILL_SECTION
    ws_ui.cell(row=R, column=1).font  = Font(bold=True)
    R += 1
    write_year_row(ws_ui, R); R += 1

    emis_links = {}
    for bi_r, lbl in {
        268: "Primary emissions (index)  [EDIT HERE]",
        269: "Fossil emissions (index)  [EDIT HERE]",
        270: "Renewables emissions (index)  [EDIT HERE]",
        271: "Secondary emissions (index)  [EDIT HERE]",
        272: "Tertiary emissions (index)  [EDIT HERE]",
        273: "Fossil sector output (index)  [EDIT HERE]",
        274: "Fossil exports (index)  [EDIT HERE]",
    }.items():
        emis_links[bi_r] = add_editable(bi_r, lbl)
    R += 1

    # ── Section 4: Public Capital ────────────────────────────────────────────
    ws_ui.cell(row=R, column=1).value = "=== SECTION 4:  PUBLIC CAPITAL BY INDUSTRY  [EDIT HERE] ==="
    ws_ui.cell(row=R, column=1).fill  = FILL_SECTION
    ws_ui.cell(row=R, column=1).font  = Font(bold=True)
    R += 1
    write_year_row(ws_ui, R); R += 1

    cap_links = {}
    for i, bi_r in enumerate(range(279, 284)):
        cap_links[bi_r] = add_editable(bi_r, f"{SECTORS[i]} public capital (index)  [EDIT HERE]")
    R += 1

    # column widths
    ws_ui.column_dimensions["A"].width = 70
    ws_ui.column_dimensions["B"].width = 5
    for y in range(BASE_YEAR, FINAL_YEAR + 1):
        ws_ui.column_dimensions[col_l(y)].width = 8

    # ── Link Baseline_Input -> User_Inputs for editable rows ─────────────────
    print("Linking Baseline_Input editable rows to User_Inputs:")
    link_to_ui(ws_bi, 238, ui_238)
    print(f"  Row 238 (GVA growth rate %)         -> User_Inputs row {ui_238}")
    for bi_r, ui_r in gva_share_links.items():
        link_to_ui(ws_bi, bi_r, ui_r)
        print(f"  Row {bi_r} ({SECTORS[bi_r-242]} VA share)      -> User_Inputs row {ui_r}")
    link_to_ui(ws_bi, 253, ui_253)
    print(f"  Row 253 (Employment growth rate %)  -> User_Inputs row {ui_253}")
    for bi_r, ui_r in emp_share_links.items():
        link_to_ui(ws_bi, bi_r, ui_r)
        print(f"  Row {bi_r} ({SECTORS[bi_r-257]} employment) -> User_Inputs row {ui_r}")
    for bi_r, ui_r in emis_links.items():
        link_to_ui(ws_bi, bi_r, ui_r)
        print(f"  Row {bi_r} ({EDITABLE_ROWS[bi_r]}) -> User_Inputs row {ui_r}")
    for bi_r, ui_r in cap_links.items():
        link_to_ui(ws_bi, bi_r, ui_r)
        print(f"  Row {bi_r} ({SECTORS[bi_r-279]} capital)  -> User_Inputs row {ui_r}")

    # Config scalars
    for (bi_r, bi_c), (ui_r, ui_c) in cfg_links.items():
        ws_bi.cell(row=bi_r, column=bi_c).value = (
            f"='{SHEET_UI}'!{get_column_letter(ui_c)}{ui_r}"
        )
    print("  Config rows 3-7 col B  -> User_Inputs config section")

    # ── Fix Baseline_Input rows 24-48: OFFSET formulas ───────────────────────
    # Col B=gY_1_1, C=gN_1_1, D=gY_2_1, E=gN_2_1, F=gY_3_1, G=gN_3_1,
    #     H=gY_4_1, I=gN_4_1, J=gY_5_1, K=gN_5_1
    # Source rows: gY->142-146, gN->163-167, col C = BASE_YEAR anchor
    col_map = [
        ("B", 142), ("C", 163),  # sector 1 (Primary)
        ("D", 143), ("E", 164),  # sector 2 (Fossil)
        ("F", 144), ("G", 165),  # sector 3 (Renewables)
        ("H", 145), ("I", 166),  # sector 4 (Secondary)
        ("J", 146), ("K", 167),  # sector 5 (Tertiary)
    ]
    for data_row in range(24, 49):
        for tgt_col, src_row in col_map:
            ws_bi.cell(row=data_row,
                       column=column_index_from_string(tgt_col)).value = (
                f"=OFFSET($C${src_row},0,$A{data_row}-{BASE_YEAR})"
            )
    print("\nBaseline_Input rows 24-48 (cols B-K): OFFSET formulas updated")

    show_only_primary_baseline_blocks(ws_bi)
    print("Baseline_Input layout updated: only VA/GVA and employment blocks remain visible")

    wb_fml.save(str(WORKBOOK))
    sanitize_excel_package(WORKBOOK)
    print(f"\nSaved  {WORKBOOK}")
    print(f"User_Inputs sheet: {R - 1} rows, {Y_END} cols")

if __name__ == "__main__":
    main()
